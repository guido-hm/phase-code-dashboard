from flask import Flask, render_template, jsonify, request
from tkinter import filedialog
import openpyxl as opxl


app = Flask(__name__)

@app.route('/api/data-civil', methods=['GET'])
def get_data_civil():
    data = {} # key = phase name. values = [estimated_qty, claimed_qty, [CATEGORIES], [TOTAL_BUDGET], [CURRENT_COSTS]]
    categories = []
    total_budgets = []
    current_costs = []

    file_name = "civil.xlsx"

    wb = opxl.load_workbook(file_name)
    ws = wb.active
    
    current_column = 'A'
    current_row = '19'

    current_cell = ws[current_column+current_row]

    possible_categories = ["Labor", "Equipment"]

    loop_one = True
    while loop_one:
        key = current_cell.value

        loop_two = True
        while loop_two:
            current_column = 'B'
            current_row = str(int(current_row)+1)
            current_cell = ws[current_column+current_row]

            if current_cell.value in possible_categories:
                categories.append(current_cell.value)

                current_column = 'F'
                current_cell = ws[current_column+current_row]
                total_budgets.append(current_cell.value)

                current_column = 'I'
                current_cell = ws[current_column+current_row]
                current_costs.append(current_cell.value)
            else:
                current_column = 'D'
                current_cell = ws[current_column+current_row]
                estimated_qty = current_cell.value

                current_column = 'G'
                current_cell = ws[current_column+current_row]
                claimed_qty = current_cell.value

                loop_two = False

        data[key] = [estimated_qty, claimed_qty, categories, total_budgets, current_costs]
        categories = []
        total_budgets = []
        current_costs = []

        current_column = 'A'
        current_row = str(int(current_row)+1)
        current_cell = ws[current_column+current_row]

        if current_cell.value[:7] != "Account":
            loop_one = False
        

    return jsonify(data)

@app.route('/api/data-mechanical', methods=['GET'])
def get_data_mechanical():
    data = {} # key = phase name. values = [estimated_qty, claimed_qty, [CATEGORIES], [TOTAL_BUDGET], [CURRENT_COSTS]]
    categories = []
    total_budgets = []
    current_costs = []


    file_name = "mechanical.xlsx"

    wb = opxl.load_workbook(file_name)
    ws = wb.active
    
    current_column = 'A'
    current_row = '19'

    current_cell = ws[current_column+current_row]

    possible_categories = ["Labor", "Equipment"]

    loop_one = True
    while loop_one:
        key = current_cell.value

        loop_two = True
        while loop_two:
            current_column = 'B'
            current_row = str(int(current_row)+1)
            current_cell = ws[current_column+current_row]

            if current_cell.value in possible_categories:
                categories.append(current_cell.value)

                current_column = 'F'
                current_cell = ws[current_column+current_row]
                total_budgets.append(current_cell.value)

                current_column = 'I'
                current_cell = ws[current_column+current_row]
                current_costs.append(current_cell.value)
            else:
                current_column = 'D'
                current_cell = ws[current_column+current_row]
                estimated_qty = current_cell.value

                current_column = 'G'
                current_cell = ws[current_column+current_row]
                claimed_qty = current_cell.value

                loop_two = False

        data[key] = [estimated_qty, claimed_qty, categories, total_budgets, current_costs]
        categories = []
        total_budgets = []
        current_costs = []

        current_column = 'A'
        current_row = str(int(current_row)+1)
        current_cell = ws[current_column+current_row]

        if current_cell.value[:7] != "Account":
            loop_one = False
        

    return jsonify(data)


@app.route('/api/data-electrical', methods=['GET'])
def get_data_electrical():
    data = {} # key = phase name. values = [estimated_qty, claimed_qty, [CATEGORIES], [TOTAL_BUDGET], [CURRENT_COSTS]]
    categories = []
    total_budgets = []
    current_costs = []


    file_name = "electrical.xlsx"

    wb = opxl.load_workbook(file_name)
    ws = wb.active
    
    current_column = 'A'
    current_row = '19'

    current_cell = ws[current_column+current_row]

    possible_categories = ["Labor", "Equipment"]

    loop_one = True
    while loop_one:
        key = current_cell.value

        loop_two = True
        while loop_two:
            current_column = 'B'
            current_row = str(int(current_row)+1)
            current_cell = ws[current_column+current_row]

            if current_cell.value in possible_categories:
                categories.append(current_cell.value)

                current_column = 'F'
                current_cell = ws[current_column+current_row]
                total_budgets.append(current_cell.value)

                current_column = 'I'
                current_cell = ws[current_column+current_row]
                current_costs.append(current_cell.value)
            else:
                current_column = 'D'
                current_cell = ws[current_column+current_row]
                estimated_qty = current_cell.value

                current_column = 'G'
                current_cell = ws[current_column+current_row]
                claimed_qty = current_cell.value

                loop_two = False

        data[key] = [estimated_qty, claimed_qty, categories, total_budgets, current_costs]
        categories = []
        total_budgets = []
        current_costs = []

        current_column = 'A'
        current_row = str(int(current_row)+1)
        current_cell = ws[current_column+current_row]

        if current_cell.value[:7] != "Account":
            loop_one = False
        

    return jsonify(data)

@app.route('/')
def home():
    return render_template('index.html')


@app.route('/civil')
def civil():
    return render_template('civil.html')

@app.route('/mechanical')
def mechanical():
    return render_template('mechanical.html')

@app.route('/electrical')
def electrical():
    return render_template('electrical.html')

if __name__ == '__main__':
    app.run(debug=False)